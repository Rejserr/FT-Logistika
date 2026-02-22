"""
Export servis za generiranje PDF i Excel izvještaja ruta.
"""
from __future__ import annotations

import io
import logging
import os
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm, mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy.orm import Session

from app.models.erp_models import NalogHeader, Partner
from app.models.routing_models import Ruta, RutaStop
from app.models.routing_order_models import NalogHeaderRutiranje
from app.models.regional_models import Regija
from app.models.mantis_models import MantisSSCC
from app.models.vehicle_models import Vozac, Vozilo

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Registracija fontova s podrškom za hrvatske znakove (č, ć, ž, đ, š)
# ---------------------------------------------------------------------------
_FONTS_DIR = r"C:\Windows\Fonts"
_fonts_registered = False


def _register_croatian_fonts() -> None:
    """Registriraj Arial font (podržava HR znakove) u ReportLab."""
    global _fonts_registered
    if _fonts_registered:
        return

    font_map = {
        "Arial": "arial.ttf",
        "Arial-Bold": "arialbd.ttf",
        "Arial-Italic": "ariali.ttf",
        "Arial-BoldItalic": "arialbi.ttf",
    }

    for name, filename in font_map.items():
        path = os.path.join(_FONTS_DIR, filename)
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont(name, path))
            except Exception as exc:
                logger.warning("Nije moguće registrirati font %s: %s", name, exc)

    _fonts_registered = True


class ExportService:
    """Servis za export ruta u PDF i Excel formate."""

    def _get_nalog(self, db: Session, nalog_uid: str) -> Any:
        """
        Dohvati nalog header - prvo u rutiranju, pa u originalu.
        Nalozi su premješteni iz nalozi_header u nalozi_header_rutiranje
        kada se stave u rutiranje.
        """
        nalog = db.get(NalogHeaderRutiranje, nalog_uid)
        if not nalog:
            nalog = db.get(NalogHeader, nalog_uid)
        return nalog

    def _get_route_regions(self, db: Session, stops, ) -> list[str]:
        """Dohvati nazive svih regija koje pokriva ruta."""
        region_ids: set[int] = set()
        for stop in stops:
            nalog = self._get_nalog(db, stop.nalog_uid)
            if nalog and nalog.regija_id:
                region_ids.add(nalog.regija_id)

        region_names: list[str] = []
        for rid in sorted(region_ids):
            regija = db.get(Regija, rid)
            if regija:
                region_names.append(regija.naziv)
        return region_names

    def _get_pallet_counts(self, db: Session, stops) -> dict[str, int]:
        """Dohvati broj paleta (distinct SSCC) po nalogu iz mantis_sscc."""
        nalog_uids = [s.nalog_uid for s in stops]
        if not nalog_uids:
            return {}

        items = (
            db.query(MantisSSCC)
            .filter(MantisSSCC.nalog_prodaje_uid.in_(nalog_uids))
            .all()
        )

        # Grupiraj po nalog_uid
        counts: dict[str, int] = {}
        grouped: dict[str, set[str]] = {}
        for it in items:
            uid = it.nalog_prodaje_uid
            if uid not in grouped:
                grouped[uid] = set()
            if it.sscc:
                grouped[uid].add(it.sscc)

        for uid in nalog_uids:
            counts[uid] = len(grouped.get(uid, set()))

        return counts

    def export_route_to_excel(self, db: Session, ruta_id: int) -> bytes:
        """
        Generiraj Excel izvještaj za rutu.

        Returns:
            bytes sadržaj Excel datoteke
        """
        ruta = db.get(Ruta, ruta_id)
        if not ruta:
            raise ValueError(f"Ruta {ruta_id} nije pronađena")

        # Dohvati povezane podatke
        vozilo = db.get(Vozilo, ruta.vozilo_id) if ruta.vozilo_id else None
        vozac = db.get(Vozac, ruta.vozac_id) if ruta.vozac_id else None

        stops = (
            db.query(RutaStop)
            .filter(RutaStop.ruta_id == ruta_id)
            .order_by(RutaStop.redoslijed)
            .all()
        )

        # Regije
        region_names = self._get_route_regions(db, stops)

        # Obrnut redoslijed za utovar (zadnji na dostavi = prvi na utovar)
        stops_utovar = list(reversed(stops))

        # Kreiraj workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Ruta {ruta_id}"

        # Stilovi
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=11)
        region_font = Font(bold=True, size=13, color="1F4E79")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Naslov
        ws["A1"] = f"Ruta #{ruta_id}"
        ws["A1"].font = header_font
        ws.merge_cells("A1:G1")

        # Informacije o ruti
        current_row = 3
        ws[f"A{current_row}"] = "Datum:"
        ws[f"B{current_row}"] = str(ruta.datum) if ruta.datum else "—"
        current_row += 1
        if ruta.raspored:
            ws[f"A{current_row}"] = "Raspored:"
            ws[f"B{current_row}"] = str(ruta.raspored)
            current_row += 1
        ws[f"A{current_row}"] = "Status:"
        ws[f"B{current_row}"] = ruta.status or "—"
        current_row += 1
        ws[f"A{current_row}"] = "Vozilo:"
        ws[f"B{current_row}"] = vozilo.oznaka if vozilo else "—"
        current_row += 1
        ws[f"A{current_row}"] = "Vozač:"
        ws[f"B{current_row}"] = f"{vozac.ime} {vozac.prezime}" if vozac else "—"
        current_row += 1
        ws[f"A{current_row}"] = "Udaljenost:"
        ws[f"B{current_row}"] = f"{float(ruta.distance_km):.1f} km" if ruta.distance_km else "—"
        current_row += 1
        ws[f"A{current_row}"] = "Trajanje:"
        ws[f"B{current_row}"] = f"{ruta.duration_min} min" if ruta.duration_min else "—"
        current_row += 1
        ws[f"A{current_row}"] = "Broj stopova:"
        ws[f"B{current_row}"] = str(len(stops))

        for row_idx in range(3, current_row + 1):
            ws[f"A{row_idx}"].font = subheader_font

        # Regije
        current_row += 2
        if region_names:
            ws[f"A{current_row}"] = f"Regija: {', '.join(region_names)}"
            ws[f"A{current_row}"].font = region_font
            ws.merge_cells(f"A{current_row}:G{current_row}")
            current_row += 1

        # Tablica stopova - Lista utovara
        current_row += 1
        ws[f"A{current_row}"] = "Lista utovara"
        ws[f"A{current_row}"].font = header_font
        ws.merge_cells(f"A{current_row}:G{current_row}")

        # Header tablice
        current_row += 2
        header_row = current_row
        headers = ["Utovar #", "Dostava #", "Partner", "Adresa", "Mjesto", "ETA", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # Podaci stopova u obrnutom redoslijedu za utovar
        for i, stop in enumerate(stops_utovar):
            row = header_row + 1 + i
            nalog = self._get_nalog(db, stop.nalog_uid)
            partner = None
            if nalog and nalog.partner_uid:
                partner = db.get(Partner, nalog.partner_uid)

            ws.cell(row=row, column=1, value=i + 1).border = thin_border
            ws.cell(row=row, column=2, value=stop.redoslijed).border = thin_border
            ws.cell(
                row=row, column=3, value=partner.naziv if partner else "—"
            ).border = thin_border
            ws.cell(
                row=row, column=4, value=partner.adresa if partner else "—"
            ).border = thin_border
            ws.cell(
                row=row, column=5, value=partner.naziv_mjesta or (partner.mjesto if partner else "—")
            ).border = thin_border
            ws.cell(
                row=row,
                column=6,
                value=stop.eta.strftime("%H:%M") if stop.eta else "—",
            ).border = thin_border
            ws.cell(row=row, column=7, value=stop.status or "—").border = thin_border

        # Širine kolona
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 15

        # Spremi u bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def export_route_to_pdf(self, db: Session, ruta_id: int) -> bytes:
        """
        Generiraj profesionalni PDF utovarni list za rutu (Landscape, HR znakovi).

        Returns:
            bytes sadržaj PDF datoteke
        """
        _register_croatian_fonts()

        FONT = "Arial"
        FONT_B = "Arial-Bold"
        FONT_I = "Arial-Italic"

        ruta = db.get(Ruta, ruta_id)
        if not ruta:
            raise ValueError(f"Ruta {ruta_id} nije pronađena")

        # Dohvati povezane podatke
        vozilo = db.get(Vozilo, ruta.vozilo_id) if ruta.vozilo_id else None
        vozac = db.get(Vozac, ruta.vozac_id) if ruta.vozac_id else None

        stops = (
            db.query(RutaStop)
            .filter(RutaStop.ruta_id == ruta_id)
            .order_by(RutaStop.redoslijed)
            .all()
        )

        # Regije
        region_names = self._get_route_regions(db, stops)

        # SSCC / pallet counts iz WMS-a
        pallet_counts = self._get_pallet_counts(db, stops)
        total_route_pallets = sum(pallet_counts.values())

        # Obrnut redoslijed za utovar (zadnji na dostavi = prvi na utovar)
        stops_utovar = list(reversed(stops))

        # ---- Landscape A4 ----
        page_w, page_h = landscape(A4)  # 842 x 595
        margin_lr = 1.2 * cm
        margin_tb = 1 * cm

        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=margin_lr,
            leftMargin=margin_lr,
            topMargin=margin_tb,
            bottomMargin=margin_tb,
        )

        usable_width = page_w - 2 * margin_lr
        elements: list[Any] = []

        # ========== ZAGLAVLJE: naslov + info u jednom redu ==========

        # Lijeva strana zaglavlja
        header_left_data = [
            [f"UTOVARNI LIST — Ruta #{ruta_id}"],
        ]
        if region_names:
            header_left_data.append([f"Regija: {', '.join(region_names)}"])
        header_left_data.append(
            [f"Generirano: {datetime.now().strftime('%d.%m.%Y %H:%M')}"]
        )

        header_left = Table(header_left_data, colWidths=[usable_width * 0.50])
        header_left_styles = [
            ("FONTNAME", (0, 0), (0, 0), FONT_B),
            ("FONTSIZE", (0, 0), (0, 0), 16),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]
        if region_names:
            header_left_styles.extend([
                ("FONTNAME", (0, 1), (0, 1), FONT_B),
                ("FONTSIZE", (0, 1), (0, 1), 12),
                ("TEXTCOLOR", (0, 1), (0, 1), colors.HexColor("#1F4E79")),
                ("FONTNAME", (0, 2), (0, 2), FONT),
                ("FONTSIZE", (0, 2), (0, 2), 8),
                ("TEXTCOLOR", (0, 2), (0, 2), colors.grey),
            ])
        else:
            header_left_styles.extend([
                ("FONTNAME", (0, 1), (0, 1), FONT),
                ("FONTSIZE", (0, 1), (0, 1), 8),
                ("TEXTCOLOR", (0, 1), (0, 1), colors.grey),
            ])
        header_left.setStyle(TableStyle(header_left_styles))

        # Desna strana zaglavlja — kompaktne info (2 stupca, manji font)
        raspored_str = str(ruta.raspored) if ruta.raspored else "—"
        datum_str = str(ruta.datum) if ruta.datum else "—"
        vozilo_str = vozilo.oznaka if vozilo else "—"
        vozac_str = f"{vozac.ime} {vozac.prezime}" if vozac else "—"
        dist_str = f"{float(ruta.distance_km):.1f} km" if ruta.distance_km else "—"
        dur_str = f"{ruta.duration_min} min" if ruta.duration_min else "—"

        paleta_str = str(total_route_pallets) if total_route_pallets > 0 else "—"
        info_rows = [
            ["Datum:", datum_str, "Vozilo:", vozilo_str],
            ["Raspored:", raspored_str, "Vozač:", vozac_str],
            ["Udaljenost:", dist_str, "Trajanje:", dur_str],
            ["Br. stopova:", str(len(stops)), "WMS Palete:", paleta_str],
        ]
        col_w = usable_width * 0.50 / 4
        info_right = Table(info_rows, colWidths=[col_w * 0.9, col_w * 1.1, col_w * 0.9, col_w * 1.1])
        info_right.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), FONT),
                    ("FONTNAME", (0, 0), (0, -1), FONT_B),
                    ("FONTNAME", (2, 0), (2, -1), FONT_B),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                    ("TOPPADDING", (0, 0), (-1, -1), 2),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LINEBELOW", (0, -1), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
                ]
            )
        )

        # Spoji u jedan red
        master_header = Table(
            [[header_left, info_right]],
            colWidths=[usable_width * 0.50, usable_width * 0.50],
        )
        master_header.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        elements.append(master_header)
        elements.append(Spacer(1, 6))

        # ========== TABLICA UTOVARA ==========

        # Header
        table_data = [
            [
                "Utovar\n#",
                "Dostava\n#",
                "Partner / Kupac",
                "Adresa",
                "Mjesto",
                "Palete",
                "ETA",
                "Potpis",
            ]
        ]

        for i, stop in enumerate(stops_utovar):
            nalog = self._get_nalog(db, stop.nalog_uid)
            partner = None
            if nalog and nalog.partner_uid:
                partner = db.get(Partner, nalog.partner_uid)

            # Partner naziv
            partner_name = "—"
            if partner:
                if partner.naziv:
                    partner_name = partner.naziv
                elif partner.ime and partner.prezime:
                    partner_name = f"{partner.ime} {partner.prezime}"

            # Adresa
            partner_adresa = partner.adresa if partner and partner.adresa else "—"

            # Mjesto
            partner_mjesto = "—"
            if partner:
                partner_mjesto = partner.naziv_mjesta or partner.mjesto or "—"

            # Palete iz WMS-a
            pc = pallet_counts.get(stop.nalog_uid, 0)
            palete_str = str(pc) if pc > 0 else "—"

            table_data.append(
                [
                    str(i + 1),
                    str(stop.redoslijed),
                    partner_name,
                    partner_adresa,
                    partner_mjesto,
                    palete_str,
                    stop.eta.strftime("%H:%M") if stop.eta else "—",
                    "",  # potpis kolona — prazno za ručni potpis
                ]
            )

        # Širine kolona za landscape (ukupno ~ usable_width)
        col_widths = [
            1.2 * cm,   # Utovar #
            1.2 * cm,   # Dostava #
            6.0 * cm,   # Partner
            6.0 * cm,   # Adresa
            3.5 * cm,   # Mjesto
            1.5 * cm,   # Palete
            1.5 * cm,   # ETA
            3.5 * cm,   # Potpis
        ]

        stops_table = Table(table_data, colWidths=col_widths, repeatRows=1)
        stops_table.setStyle(
            TableStyle(
                [
                    # Header
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), FONT_B),
                    ("FONTSIZE", (0, 0), (-1, 0), 8),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
                    # Data rows
                    ("FONTNAME", (0, 1), (-1, -1), FONT),
                    ("FONTSIZE", (0, 1), (-1, -1), 9),
                    ("ALIGN", (0, 1), (1, -1), "CENTER"),  # Utovar# i Dostava# centrirano
                    ("ALIGN", (5, 1), (5, -1), "CENTER"),  # Palete centrirano
                    ("ALIGN", (6, 1), (6, -1), "CENTER"),  # ETA centrirano
                    ("VALIGN", (0, 1), (-1, -1), "MIDDLE"),
                    # Grid
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#999999")),
                    ("LINEBELOW", (0, 0), (-1, 0), 1.5, colors.HexColor("#1F4E79")),
                    # Padding
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                    ("TOPPADDING", (0, 0), (-1, 0), 6),
                    ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
                    ("TOPPADDING", (0, 1), (-1, -1), 5),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    # Alternating rows
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F6FA")]),
                ]
            )
        )
        elements.append(stops_table)

        # ========== FOOTER NOTE ==========
        elements.append(Spacer(1, 8))
        note_style = ParagraphStyle(
            "FooterNote",
            fontName=FONT_I,
            fontSize=7,
            textColor=colors.HexColor("#888888"),
        )
        elements.append(
            Paragraph(
                "Redoslijed utovara: zadnja dostava se utovaruje prva. "
                "Kolona 'Potpis' služi za potvrdu preuzimanja robe.",
                note_style,
            )
        )

        # Build PDF
        doc.build(elements)
        output.seek(0)
        return output.getvalue()


# Singleton
export_service = ExportService()
