"""
Servis za import artikl-kriterija veza iz XLSX datoteke.

Očekivani XLSX format (header u prvom redu):
  artikl | kriterija
  12345  | Raster
  67890  | Raster
"""
from __future__ import annotations

import logging
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.erp_models import Artikl, ArtiklKriterija, KriterijaSku

logger = logging.getLogger(__name__)


def _normalize(val: Any) -> str:
    """Normalize header name."""
    return str(val).strip().lower().replace(" ", "_")


def import_artikl_kriterija_from_xlsx(
    file_content: bytes,
    db: Session,
) -> dict[str, Any]:
    """Import artikl-kriterija veza iz XLSX datoteke.

    Returns dict with keys: imported, skipped, errors.
    """
    import io

    wb = load_workbook(filename=io.BytesIO(file_content), read_only=True)
    ws = wb.active
    if ws is None:
        return {"imported": 0, "skipped": 0, "errors": ["XLSX nema aktivni sheet."]}

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {"imported": 0, "skipped": 0, "errors": ["XLSX nema podataka (samo header ili prazno)."]}

    # Parse header
    header = [_normalize(cell) if cell else "" for cell in rows[0]]

    # Map column names
    artikl_idx = None
    kriterija_idx = None

    artikl_aliases = {"artikl", "sifra", "šifra", "artikl_sifra", "sifra_artikla"}
    kriterija_aliases = {"kriterija", "kriterij", "criteria", "criterion", "tip"}

    for i, h in enumerate(header):
        if h in artikl_aliases:
            artikl_idx = i
        if h in kriterija_aliases:
            kriterija_idx = i

    if artikl_idx is None:
        return {"imported": 0, "skipped": 0, "errors": ["Kolona 'artikl' ili 'sifra' nije pronađena u headeru."]}
    if kriterija_idx is None:
        return {"imported": 0, "skipped": 0, "errors": ["Kolona 'kriterija' nije pronađena u headeru."]}

    # Cache: kriterija_naziv -> KriterijaSku object
    kriterija_cache: dict[str, KriterijaSku] = {}
    existing_kriterije = db.execute(select(KriterijaSku)).scalars().all()
    for k in existing_kriterije:
        kriterija_cache[k.naziv.strip().lower()] = k

    # Cache: artikl sifra -> Artikl object (for naziv lookup)
    artikl_cache: dict[str, Artikl] = {}

    imported = 0
    skipped = 0
    errors: list[str] = []

    for row_num, row in enumerate(rows[1:], start=2):
        try:
            raw_artikl = row[artikl_idx]
            raw_kriterija = row[kriterija_idx]

            if not raw_artikl or not raw_kriterija:
                skipped += 1
                continue

            artikl_sifra = str(raw_artikl).strip()
            kriterija_naziv = str(raw_kriterija).strip()

            # Resolve kriterija - create if doesn't exist
            key = kriterija_naziv.lower()
            if key not in kriterija_cache:
                new_k = KriterijaSku(naziv=kriterija_naziv)
                db.add(new_k)
                db.flush()
                kriterija_cache[key] = new_k
                logger.info(f"Automatski kreiran kriterij: {kriterija_naziv}")

            kriterija_obj = kriterija_cache[key]

            # Resolve artikl naziv from artikli table
            artikl_naziv = None
            if artikl_sifra not in artikl_cache:
                result = db.execute(
                    select(Artikl).where(Artikl.artikl == artikl_sifra)
                ).scalar_one_or_none()
                if result:
                    artikl_cache[artikl_sifra] = result
            if artikl_sifra in artikl_cache:
                artikl_naziv = artikl_cache[artikl_sifra].naziv

            # Check for duplicate
            existing = db.execute(
                select(ArtiklKriterija).where(
                    ArtiklKriterija.artikl == artikl_sifra,
                    ArtiklKriterija.kriterija_id == kriterija_obj.id,
                )
            ).scalar_one_or_none()

            if existing:
                skipped += 1
                continue

            new_ak = ArtiklKriterija(
                artikl=artikl_sifra,
                artikl_naziv=artikl_naziv,
                kriterija_id=kriterija_obj.id,
            )
            db.add(new_ak)
            imported += 1

        except Exception as e:
            errors.append(f"Red {row_num}: {str(e)}")
            logger.warning(f"Greška u redu {row_num}: {e}")

    db.commit()
    wb.close()

    return {
        "imported": imported,
        "skipped": skipped,
        "errors": errors,
    }
