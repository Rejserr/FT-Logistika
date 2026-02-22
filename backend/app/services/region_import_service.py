"""
Servis za import regija i poštanskih brojeva iz CSV ili XLSX.

Očekivani format datoteke (header u prvom redu):
  Postanski_broj, Mjesto, regija
  ili
  postanski_broj, mjesto, regija

Primjer:
  Postanski_broj;Mjesto;regija
  10000;Sljeme;ZAGREBAČKA
  10000;Zagreb;ZAGREBAČKA
  10010;Buzin;ZAGREBAČKA
"""
from __future__ import annotations

import csv
import io
import logging
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.regional_models import PostanskiBroj, Regija

logger = logging.getLogger(__name__)

# Moguća imena kolona (normalizirana: lower, strip, zamjena razmaka s _)
COL_POSTANSKI = ("postanski_broj", "postanskibroj", "postanski broj", "pb", "poštanski_broj")
COL_MJESTO = ("mjesto", "grad", "naziv_mjesta", "naziv mjesta", "place")
COL_REGIJA = ("regija", "region", "naziv_regije", "naziv regije")


def _normalize_header(name: str) -> str:
    if not name or not isinstance(name, str):
        return ""
    return name.strip().lower().replace(" ", "_").replace("\ufeff", "")


def _find_column_index(headers: list[str], aliases: tuple[str, ...]) -> int | None:
    for i, h in enumerate(headers):
        n = _normalize_header(h)
        if n in aliases:
            return i
    return None


def _parse_csv(content: bytes) -> list[dict[str, str]]:
    """Parsiraj CSV; podržava tab, ; ili , kao separator."""
    text = content.decode("utf-8-sig")
    rows: list[list[str]] = []
    for delim in ("\t", ";", ","):
        reader = csv.reader(io.StringIO(text), delimiter=delim)
        rows = list(reader)
        if not rows:
            return []
        if len(rows[0]) >= 2:
            break

    first = rows[0]

    idx_pb = _find_column_index([str(h) for h in first], COL_POSTANSKI)
    idx_mjesto = _find_column_index([str(h) for h in first], COL_MJESTO)
    idx_regija = _find_column_index([str(h) for h in first], COL_REGIJA)

    if idx_pb is None or idx_regija is None:
        raise ValueError(
            "Datoteka mora imati kolone za poštanski broj i regiju. "
            "Očekivani nazivi: Postanski_broj (ili postanski_broj), Mjesto, regija."
        )

    result = []
    for row in rows[1:]:
        if len(row) <= max(idx_pb, idx_regija or 0):
            continue
        postanski = str(row[idx_pb]).strip() if idx_pb is not None else ""
        mjesto = str(row[idx_mjesto]).strip() if idx_mjesto is not None and idx_mjesto < len(row) else ""
        regija_naziv = str(row[idx_regija]).strip() if idx_regija is not None else ""
        if not postanski or not regija_naziv:
            continue
        result.append({"postanski_broj": postanski, "mjesto": mjesto, "regija": regija_naziv})
    return result


def _parse_xlsx(content: bytes) -> list[dict[str, str]]:
    """Parsiraj XLSX; prvi sheet, prvi red = header."""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if not ws:
        raise ValueError("Excel datoteka nema aktivni sheet.")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    first = [str(c) if c is not None else "" for c in rows[0]]
    idx_pb = _find_column_index(first, COL_POSTANSKI)
    idx_mjesto = _find_column_index(first, COL_MJESTO)
    idx_regija = _find_column_index(first, COL_REGIJA)

    if idx_pb is None or idx_regija is None:
        raise ValueError(
            "Datoteka mora imati kolone za poštanski broj i regiju. "
            "Očekivani nazivi: Postanski_broj (ili postanski_broj), Mjesto, regija."
        )

    result = []
    for row in rows[1:]:
        row = list(row) if row else []
        if len(row) <= max(idx_pb, idx_regija or 0):
            continue
        postanski = str(row[idx_pb] or "").strip()
        mjesto = str(row[idx_mjesto] or "").strip() if idx_mjesto is not None and idx_mjesto < len(row) else ""
        regija_naziv = str(row[idx_regija] or "").strip()
        if not postanski or not regija_naziv:
            continue
        result.append({"postanski_broj": postanski, "mjesto": mjesto, "regija": regija_naziv})
    return result


def import_regije_i_postanski(
    db: Session, content: bytes, filename: str
) -> dict[str, Any]:
    """
    Uvezi regije i poštanske brojeve iz CSV ili XLSX.

    Returns:
        dict s keys: regije_created, regije_existing, postanski_created, postanski_updated, errors
    """
    filename_lower = filename.lower()
    if filename_lower.endswith(".csv"):
        rows = _parse_csv(content)
    elif filename_lower.endswith(".xlsx") or filename_lower.endswith(".xls"):
        rows = _parse_xlsx(content)
    else:
        raise ValueError("Podržani formati: .csv, .xlsx")

    if not rows:
        return {
            "regije_created": 0,
            "regije_existing": 0,
            "postanski_created": 0,
            "postanski_updated": 0,
            "rows_processed": 0,
            "errors": ["Nema valjanih redaka u datoteci."],
        }

    regije_by_naziv: dict[str, Regija] = {}
    regije_created = 0
    regije_existing = 0
    postanski_created = 0
    postanski_updated = 0
    errors: list[str] = []

    for row in rows:
        regija_naziv = row["regija"].strip()
        postanski_broj = row["postanski_broj"].strip()
        mjesto = (row.get("mjesto") or "").strip()

        if not regija_naziv or not postanski_broj:
            continue

        # Get or create Regija (match po nazivu, case-insensitive)
        regija_key = regija_naziv.upper()
        if regija_key not in regije_by_naziv:
            existing = db.execute(
                select(Regija).where(Regija.naziv.ilike(regija_naziv)).limit(1)
            ).scalars().first()
            if existing:
                regije_by_naziv[regija_key] = existing
                regije_existing += 1
            else:
                regija = Regija(naziv=regija_naziv, aktivan=True)
                db.add(regija)
                db.flush()
                regije_by_naziv[regija_key] = regija
                regije_created += 1

        regija = regije_by_naziv[regija_key]

        # Upsert PostanskiBroj: (postanski_broj, naziv_mjesta) unique
        existing_pb = db.execute(
            select(PostanskiBroj).where(
                PostanskiBroj.postanski_broj == postanski_broj,
                PostanskiBroj.naziv_mjesta == mjesto,
            )
        ).scalars().first()
        if existing_pb:
            existing_pb.regija_id = regija.id
            postanski_updated += 1
        else:
            db.add(
                PostanskiBroj(
                    postanski_broj=postanski_broj,
                    naziv_mjesta=mjesto,
                    regija_id=regija.id,
                )
            )
            postanski_created += 1

    db.commit()

    return {
        "regije_created": regije_created,
        "regije_existing": regije_existing,
        "postanski_created": postanski_created,
        "postanski_updated": postanski_updated,
        "rows_processed": len(rows),
        "errors": errors,
    }
